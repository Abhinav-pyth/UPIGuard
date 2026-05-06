import urllib.request, json
from pathlib import Path

BASE = 'http://localhost:8000'

def post(path, body, token=None):
    data = json.dumps(body).encode()
    headers = {'Content-Type': 'application/json'}
    if token:
        headers['Authorization'] = f'Bearer {token}'
    req = urllib.request.Request(f'{BASE}{path}', data=data, headers=headers, method='POST')
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, json.loads(r.read())
    except urllib.error.HTTPError as e:
        return e.code, json.loads(e.read())

def get(path, token=None):
    req = urllib.request.Request(f'{BASE}{path}')
    if token:
        req.add_header('Authorization', f'Bearer {token}')
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, json.loads(r.read())
    except urllib.error.HTTPError as e:
        return e.code, json.loads(e.read())

print("=== 1. Health ===")
s, d = get('/api/health')
print(f"  {s} records={d.get('breach_records_loaded')} unique={d.get('unique_upi_ids_in_db')}")

print("\n=== 2. Register ===")
s, d = post('/api/register', {'email': 'smoke@upiguard.dev', 'password': 'SecurePass123', 'consent_given': True})
print(f"  {s}", d.get('message', d))

print("\n=== 3. OTP Request ===")
s, d = post('/api/otp/request', {'email': 'smoke@upiguard.dev'})
otp = d.get('otp_dev', '')
print(f"  {s} otp_dev={otp}")

print("\n=== 4. OTP Verify ===")
s, d = post('/api/otp/verify', {'email': 'smoke@upiguard.dev', 'code': otp})
token = d.get('access_token', '')
print(f"  {s} token={'OK (' + token[:20] + '...)' if token else d}")

print("\n=== 5a. Check WITHOUT auth (expect 401) ===")
s, d = post('/api/check', {'upi_id': 'rahul@oksbi'})
print(f"  {s} detail={d.get('detail')}")

print("\n=== 5b. Check WITH auth ===")
s, d = post('/api/check', {'upi_id': 'testuser@paytm'}, token=token)
print(f"  {s} is_compromised={d.get('is_compromised')} breaches={d.get('breach_count')}")

print("\n=== 6. Dashboard ===")
s, d = get('/api/dashboard', token=token)
print(f"  {s} email={d.get('email')} total_checks={d.get('total_checks')}")

print("\n=== 7. Admin spreadsheet ===")
xlsx = Path('data/upi_admin_log.xlsx')
print(f"  File exists: {xlsx.exists()}")
if xlsx.exists():
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    print(f"  Rows (incl header): {len(rows)}")
    if len(rows) > 1:
        print(f"  Last row: {rows[-1]}")

print("\n=== ALL DONE ===")
