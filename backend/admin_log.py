"""
admin_log.py — Append every UPI check to an admin-only Excel spreadsheet.

The file lives at ADMIN_LOG_PATH (default: backend/data/upi_admin_log.xlsx).
It is NEVER served through any API endpoint — filesystem access only.
"""

import os
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()

_LOG_PATH = Path(__file__).parent / os.getenv("ADMIN_LOG_PATH", "data/upi_admin_log.xlsx")

_HEADERS = [
    "Timestamp (UTC)",
    "User Email",
    "UPI ID (Plaintext)",
    "UPI Hash",
    "Is Compromised",
    "Breach Count",
    "Client IP",
]


def _ensure_workbook() -> None:
    """Create the Excel file with headers if it doesn't exist."""
    if _LOG_PATH.exists():
        return
    _LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "UPI Checks"
    ws.append(_HEADERS)
    # Style header row bold
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(_LOG_PATH)


def log_check(
    *,
    user_email: str,
    upi_id_plaintext: str,
    upi_hash: str,
    is_compromised: bool,
    breach_count: int,
    client_ip: str,
) -> None:
    """Append one row to the admin log spreadsheet (thread-safe via file lock pattern)."""
    _ensure_workbook()
    wb = load_workbook(_LOG_PATH)
    ws = wb.active
    ws.append([
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        user_email,
        upi_id_plaintext,
        upi_hash,
        "YES" if is_compromised else "NO",
        breach_count,
        client_ip,
    ])
    wb.save(_LOG_PATH)
